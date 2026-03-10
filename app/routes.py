import os
import io
import csv
import uuid
import time
import zipfile
import threading
_log_lock = threading.Lock()
from datetime import datetime
from flask import (
    Blueprint, render_template, request,
    send_file, redirect, url_for, flash, current_app, jsonify
)
from werkzeug.utils import secure_filename

try:
    from app.pdf_processor import add_folios
    PDF_PROCESSOR_AVAILABLE = True
except ImportError as e:
    print(f"[ERROR] Could not import pdf_processor: {e}")
    PDF_PROCESSOR_AVAILABLE = False

try:
    from pdf2image import convert_from_bytes
    PDF_PREVIEW_AVAILABLE = True
except ImportError:
    print("[WARNING] pdf2image not available. Preview disabled.")
    PDF_PREVIEW_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    print("[WARNING] openpyxl not available. Excel export disabled.")
    XLSX_AVAILABLE = False


main = Blueprint("main", __name__)


def get_temp_folder():
    return current_app.config["TEMP_FOLDER"]


def get_log_path():
    return os.path.join(current_app.config["LOGS_FOLDER"], "folios.txt")


def cleanup_temp_files(temp_folder, max_age_minutes=120):
    if not os.path.exists(temp_folder):
        return
    now = time.time()
    cutoff = now - (max_age_minutes * 60)
    for filename in os.listdir(temp_folder):
        path = os.path.join(temp_folder, filename)
        try:
            if os.path.isfile(path) and os.path.getmtime(path) < cutoff:
                os.remove(path)
        except OSError:
            pass

def is_valid_pdf(path):
    try:
        with open(path, "rb") as f:
            return f.read(4) == b"%PDF"
    except OSError:
        return False

def parse_form_params(form, suffix=""):
    def field(name):
        return form.get(f"{name}{suffix}")

    start_number = int(field("start_number") or 1)
    start_page   = int(field("start_page") or 1)
    font_size    = int(field("font_size") or 14)
    offset_cm    = float(field("offset") or 1.0)
    corner       = field("corner") or "bottom-right"
    orientation  = field("orientation") or "horizontal"

    end_page_raw = field("end_page")
    end_page = int(end_page_raw) if end_page_raw and end_page_raw.isdigit() else None

    return {
        "start_number": start_number,
        "start_page":   start_page,
        "end_page":     end_page,
        "font_size":    font_size,
        "offset_cm":    offset_cm,
        "corner":       corner,
        "orientation":  orientation,
    }

def read_log_entries():
    log_path = get_log_path()
    entries = []
    if not os.path.exists(log_path):
        return entries

    with open(log_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for i, line in enumerate(reversed(lines)):
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(" | ")]

        def get(idx, prefix=""):
            if idx < len(parts):
                return parts[idx].replace(prefix, "").strip()
            return "—"

        entries.append({
            "index":    len(lines) - 1 - i,
            "date":     get(0),
            "status":   get(1),
            "folios":   get(2),
            "pages":    get(3),
            "corner":   get(4, "Corner: "),
            "filename": get(5, "File: "),
            "batch":    get(6, "Batch: "),
            "duration": get(7, "Duration: "),
            "ip":       get(8, "IP: "),
        })
    return entries


def apply_filters(entries, args):
    status   = args.get("status", "").strip().upper()
    filename = args.get("filename", "").strip().lower()
    date_from = args.get("date_from", "").strip()
    date_to   = args.get("date_to", "").strip()

    result = []
    for e in entries:
        if status and e["status"].strip().upper() != status:
            continue
        if filename and filename not in e["filename"].lower():
            continue
        if date_from:
            try:
                if e["date"][:10] < date_from:
                    continue
            except Exception:
                pass
        if date_to:
            try:
                if e["date"][:10] > date_to:
                    continue
            except Exception:
                pass
        result.append(e)
    return result


@main.route("/", methods=["GET", "POST"])
def index():
    cleanup_temp_files(current_app.config["TEMP_FOLDER"])

    if request.method == "POST":
        file = request.files.get("pdf_file")

        if not file or file.filename == "":
            flash("No se seleccionó ningún archivo.", "error")
            return redirect(url_for("main.index"))

        if not file.filename.lower().endswith(".pdf"):
            flash("El archivo debe ser un PDF.", "error")
            return redirect(url_for("main.index"))

        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        if file_size == 0:
            flash("El archivo está vacío.", "error")
            return redirect(url_for("main.index"))

        try:
            file_id   = str(uuid.uuid4())
            safe_name = secure_filename(file.filename) or f"documento_{uuid.uuid4().hex[:8]}.pdf"
            safe_name = safe_name[:100]
            temp_in   = os.path.join(get_temp_folder(), f"{file_id}_in.pdf")
            temp_out  = os.path.join(get_temp_folder(), f"{file_id}_out.pdf")

            with open(temp_in, "wb") as f:
                while chunk := file.read(8192):
                    f.write(chunk)

            params    = parse_form_params(request.form)
            client_ip = request.headers.get("X-Forwarded-For", request.remote_addr)

            success = add_folios(
                input_path=temp_in,
                output_path=temp_out,
                preview_mode=False,
                log_folder=current_app.config["LOGS_FOLDER"],
                filename=safe_name,
                client_ip=client_ip,
                **params,
            )

            if success and os.path.exists(temp_out):
                return send_file(
                    temp_out,
                    mimetype="application/pdf",
                    as_attachment=True,
                    download_name=f"Foliado_{safe_name}",
                )

            flash("Error al procesar el documento. Verifica que el rango de páginas sea válido.", "error")

        except Exception as e:
            flash(f"Error inesperado: {str(e)}", "error")

        return redirect(url_for("main.index"))

    return render_template(
        "index.html",
        pdf_processor_available=PDF_PROCESSOR_AVAILABLE,
        pdf_preview_available=PDF_PREVIEW_AVAILABLE,
    )


@main.route("/preview", methods=["POST"])
def preview():
    if not PDF_PREVIEW_AVAILABLE:
        return "Vist previa no disponible.", 501

    file = request.files.get("pdf_file")
    if not file:
        return "No se recibió ningún archivo.", 400

    file.seek(0, os.SEEK_END)
    size_mb = file.tell() / (1024 * 1024)
    file.seek(0)
    if size_mb > 30:
        return "Archivo demasiado grande para vista previa (máximo 30MB)", 413

    try:
        file_id  = str(uuid.uuid4())
        temp_in  = os.path.join(get_temp_folder(), f"prev_in_{file_id}.pdf")
        temp_out = os.path.join(get_temp_folder(), f"prev_out_{file_id}.pdf")

        file.save(temp_in)

        if not is_valid_pdf(temp_in):
            return "El archivo no es un PDF válido.", 400

        # Sanitizar si es necesario
        try:
            from pypdf import PdfReader
            PdfReader(temp_in)
        except Exception:
            from app.pdf_processor import sanitize_pdf
            sanitized = temp_in.replace(".pdf", "_sanitized.pdf")
            if sanitize_pdf(temp_in, sanitized):
                temp_in = sanitized

        params = parse_form_params(request.form, suffix="_prev")
        params["end_page"] = params["start_page"]

        add_folios(
            input_path=temp_in,
            output_path=temp_out,
            preview_mode=True,
            log_folder=current_app.config["LOGS_FOLDER"],
            **params,
        )

        with open(temp_out, "rb") as f:
            images = convert_from_bytes(
                f.read(), first_page=1, last_page=1, fmt="png", dpi=72
            )

        img_io = io.BytesIO()
        images[0].save(img_io, format="PNG")
        img_io.seek(0)
        return send_file(img_io, mimetype="image/png")

    except Exception as e:
        return f"Error al generar vista previa: {str(e)}", 500


@main.route("/instructions")
def instructions():
    return render_template("instructions.html")


@main.route("/history")
def history():
    entries   = read_log_entries()
    filtered  = apply_filters(entries, request.args)

    total_pages = sum(
        int(e["pages"].replace("Pages: ", "").strip())
        for e in entries
        if e["pages"].replace("Pages: ", "").strip().isdigit()
    )

    # Paginación
    PER_PAGE = 50
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1

    total_filtered = len(filtered)
    total_pages_nav = (total_filtered + PER_PAGE - 1) // PER_PAGE
    page = min(page, max(1, total_pages_nav))
    start = (page - 1) * PER_PAGE
    paginated = filtered[start:start + PER_PAGE]

    return render_template(
        "history.html",
        entries=paginated,
        total_entries=len(entries),
        total_pages=total_pages,
        filters=request.args,
        page=page,
        total_pages_nav=total_pages_nav,
        total_filtered=total_filtered,
    )


@main.route("/history/delete/<int:line_index>", methods=["POST"])
def history_delete_entry(line_index):
    log_path = get_log_path()
    if not os.path.exists(log_path):
        return jsonify({"ok": False, "error": "Log not found"}), 404

    with _log_lock:
        with open(log_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        if line_index < 0 or line_index >= len(lines):
            return jsonify({"ok": False, "error": "Index out of range"}), 400

        lines.pop(line_index)

        with open(log_path, "w", encoding="utf-8") as f:
            f.writelines(lines)

    return jsonify({"ok": True})


@main.route("/history/delete-all", methods=["POST"])
def history_delete_all():
    log_path = get_log_path()
    with _log_lock:
        if os.path.exists(log_path):
            with open(log_path, "w", encoding="utf-8") as f:
                f.write("")
    return redirect(url_for("main.history"))


@main.route("/history/export")
def history_export():
    fmt      = request.args.get("format", "csv").lower()
    entries  = read_log_entries()
    filtered = apply_filters(entries, request.args)

    headers = ["Fecha", "Estado", "Folios", "Páginas", "Esquina", "Archivo", "Lote", "Duración (s)", "IP"]

    def row(e):
        return [
            e["date"], e["status"], e["folios"], e["pages"],
            e["corner"], e["filename"], e["batch"] or "—",
            e["duration"], e["ip"],
        ]

    if fmt == "xlsx":
        if not XLSX_AVAILABLE:
            flash("openpyxl no está instalado. Ejecuta: pip install openpyxl", "error")
            return redirect(url_for("main.history"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial"

        header_font  = Font(bold=True, color="000000")
        header_fill  = PatternFill("solid", fgColor="C8A020")
        header_align = Alignment(horizontal="center")

        ws.append(headers)
        for cell in ws[1]:
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = header_align

        for e in filtered:
            ws.append(row(e))

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"historial_folios_{datetime.now().strftime('%Y%m%d')}.xlsx",
        )

    # CSV (default)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for e in filtered:
        writer.writerow(row(e))

    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"historial_folios_{datetime.now().strftime('%Y%m%d')}.csv",
    )


@main.route("/foliar-multiple", methods=["POST"])
def foliar_multiple():
    files = request.files.getlist("pdf_files")

    if not files or all(f.filename == "" for f in files):
        return "No se recibieron archivos.", 400

    params      = parse_form_params(request.form)
    temp_folder = get_temp_folder()
    log_folder  = current_app.config["LOGS_FOLDER"]
    client_ip   = request.headers.get("X-Forwarded-For", request.remote_addr)

    zip_buffer = io.BytesIO()
    errores    = []
    current_number = params["start_number"]

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for file in files:
            if not file.filename.lower().endswith(".pdf"):
                errores.append(file.filename)
                continue

            file_id   = str(uuid.uuid4())
            safe_name = secure_filename(file.filename) or f"documento_{uuid.uuid4().hex[:8]}.pdf"
            safe_name = safe_name[:100]
            temp_in   = os.path.join(temp_folder, f"{file_id}_in.pdf")
            temp_out  = os.path.join(temp_folder, f"{file_id}_out.pdf")

            try:
                with open(temp_in, "wb") as f:
                    while chunk := file.read(8192):
                        f.write(chunk)

                if not is_valid_pdf(temp_in):
                    errores.append(safe_name)
                    continue

                from pypdf import PdfReader
                from app.pdf_processor import sanitize_pdf
                try:
                    reader = PdfReader(temp_in)
                except Exception:
                    sanitized = temp_in.replace(".pdf", "_sanitized.pdf")
                    if sanitize_pdf(temp_in, sanitized):
                        reader = PdfReader(sanitized)
                    else:
                        errores.append(safe_name)
                        continue

                start_idx      = max(0, params["start_page"] - 1)
                end_idx        = min(len(reader.pages), params["end_page"] if params["end_page"] else len(reader.pages))
                pages_to_folio = end_idx - start_idx

                file_params = dict(params)
                file_params["start_number"] = current_number

                success = add_folios(
                    input_path=temp_in,
                    output_path=temp_out,
                    preview_mode=False,
                    log_folder=log_folder,
                    filename=safe_name,
                    batch=len(files),
                    client_ip=client_ip,
                    **file_params,
                )

                if success and os.path.exists(temp_out):
                    file_number = str(files.index(file) + 1).zfill(3)
                    zf.write(temp_out, f"{file_number}_Foliado_{safe_name}")
                    current_number += pages_to_folio
                else:
                    errores.append(safe_name)

            except Exception as e:
                errores.append(safe_name)
                print(f"[ERROR] {safe_name}: {e}")

            finally:
                for path in [temp_in, temp_out]:
                    if os.path.exists(path):
                        try:
                            os.remove(path)
                        except OSError:
                            pass

        if errores:
            error_content = "Los siguientes archivos no pudieron ser foliados:\n\n"
            error_content += "\n".join(f"- {e}" for e in errores)
            zf.writestr("ERRORES.txt", error_content)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name="Foliados.zip",
    )

@main.route("/page-count", methods=["POST"])
def page_count():
    file = request.files.get("pdf_file")
    if not file:
        return jsonify({"count": None}), 400

    try:
        file_id  = str(uuid.uuid4())
        temp_in  = os.path.join(get_temp_folder(), f"pc_{file_id}.pdf")
        file.save(temp_in)

        try:
            from pypdf import PdfReader
            reader = PdfReader(temp_in)
            count  = len(reader.pages)
        except Exception:
            from app.pdf_processor import sanitize_pdf
            sanitized = temp_in.replace(".pdf", "_s.pdf")
            if sanitize_pdf(temp_in, sanitized):
                from pypdf import PdfReader
                reader = PdfReader(sanitized)
                count  = len(reader.pages)
                os.remove(sanitized)
            else:
                count = None

        os.remove(temp_in)
        return jsonify({"count": count})

    except Exception as e:
        return jsonify({"count": None, "error": str(e)}), 500

@main.app_errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404


@main.app_errorhandler(500)
def internal_server_error(e):
    return render_template("500.html"), 500